import io
from datetime import datetime, date, timedelta
import openpyxl
import streamlit as st
import pandas as pd
from openpyxl.styles import Border, Side, Font

st.set_page_config(page_title="転記用生成ツール", layout="wide")

FMT_PATH = "転記用FMT.xlsx"

# ==============================
# 罫線
# ==============================
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ==============================
# 日付変換
# ==============================
def to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y/%m/%d").date()
    except:
        return None


# ==============================
# 転記ツール（完全そのまま）
# ==============================
def parse_input(ws):
    media_blocks = {}
    current_media = None

    for row in range(1, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value

        a_date = to_date(a)
        b_date = to_date(b)

        if isinstance(a, str) and not a_date:
            current_media = a.strip()
            if current_media not in media_blocks:
                media_blocks[current_media] = []
            continue

        if current_media and a_date and b_date and c:
            media_blocks[current_media].append({
                "start": a_date,
                "end": b_date,
                "name": str(c).strip()
            })

    return media_blocks


def safe_sheet_title(title, used_titles):
    """Excelのシート名制限に合わせて、重複しないシート名を作る。"""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    safe_title = str(title)
    for ch in invalid_chars:
        safe_title = safe_title.replace(ch, '_')

    safe_title = safe_title[:31] or 'Sheet'
    base = safe_title
    idx = 2

    while safe_title in used_titles:
        suffix = f'_{idx}'
        safe_title = base[:31 - len(suffix)] + suffix
        idx += 1

    used_titles.add(safe_title)
    return safe_title


def merge_media_blocks(base_blocks, add_blocks):
    """複数の元データシートから取得した媒体ブロックを、媒体名ごとに合算する。"""
    for media, records in add_blocks.items():
        if media not in base_blocks:
            base_blocks[media] = []
        base_blocks[media].extend(records)
    return base_blocks


def build_workbook(input_bytes, fmt_bytes, selected_sheets):
    """
    転記用エクセルを生成する。

    変更点：
    - 元データ側は複数シート選択可能
    - 選択した複数シートの内容を媒体名ごとに統合
    - 出力側は従来どおり媒体ごとにシート分け
    """
    src_wb = openpyxl.load_workbook(io.BytesIO(input_bytes), data_only=True)

    if isinstance(selected_sheets, str):
        selected_sheets = [selected_sheets]

    out_wb = openpyxl.load_workbook(io.BytesIO(fmt_bytes))
    template_ws = out_wb.worksheets[0]
    out_wb.remove(template_ws)

    media_blocks = {}

    for sheet_name in selected_sheets:
        if sheet_name not in src_wb.sheetnames:
            continue

        src_ws = src_wb[sheet_name]
        sheet_blocks = parse_input(src_ws)
        merge_media_blocks(media_blocks, sheet_blocks)

    total_days = 0
    sheet_count = 0
    used_titles = set()

    for media, records in media_blocks.items():

        if not records:
            continue

        ws = out_wb.copy_worksheet(template_ws)
        ws.title = safe_sheet_title(media, used_titles)
        sheet_count += 1

        campaigns = sorted(set(r["name"] for r in records))

        for i, name in enumerate(campaigns):
            ws.cell(row=1, column=7 + i).value = name

        min_date = min(r["start"] for r in records)
        max_date = max(r["end"] for r in records)

        current = min_date
        row = 2

        while current <= max_date:
            ws.cell(row=row, column=1).value = current
            ws.cell(row=row, column=1).number_format = "yyyy/mm/dd"

            for i, name in enumerate(campaigns):
                flag = 0
                for r in records:
                    if r["name"] == name and r["start"] <= current <= r["end"]:
                        flag = 1
                        break

                ws.cell(row=row, column=7 + i).value = flag

            current += timedelta(days=1)
            row += 1
            total_days += 1

    if sheet_count == 0:
        ws = out_wb.create_sheet("NoData")
        ws.cell(1, 1).value = "データを取得できませんでした"

    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)

    return output.getvalue(), sheet_count, total_days


# ==============================
# リッジ回帰による施策評価
# ==============================
import numpy as np
from sklearn.linear_model import RidgeCV
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import make_pipeline
from sklearn.metrics import mean_squared_error, r2_score
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

TARGET_COL_NAME = "cv"          # 目的変数。通常はCV
FEATURE_START_COL_IDX = 6       # G列以降を施策フラグとして扱う（0始まりなので6）
RIDGE_ALPHAS = np.logspace(-3, 3, 30)


def safe_float(v):
    try:
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def make_unique_sheet_name(wb, base_name):
    """Excelの31文字制限と重複回避。"""
    base = str(base_name)[:31] if base_name else "Sheet"
    if base not in wb.sheetnames:
        return base

    for i in range(1, 100):
        suffix = f"_{i}"
        name = base[:31 - len(suffix)] + suffix
        if name not in wb.sheetnames:
            return name

    return base[:28] + "_xx"


def auto_fit_columns(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def confidence_label(on_days, total_days, abs_std_coef, corr_max):
    """
    施策評価の信頼度。
    - ON日数が少なすぎる/多すぎるものは評価しづらい
    - 他施策との相関が高いものは単独効果としては注意
    """
    if total_days == 0:
        return "🔴低", "評価対象データがありません"

    on_rate = on_days / total_days

    if on_days < 3:
        return "🔴低", "ON日数が少なすぎるため参考値です"
    if on_rate >= 0.95:
        return "🔴低", "ほぼ常時ONのため効果分離が困難です"
    if on_rate <= 0.05:
        return "🔴低", "実施日が少ないため効果分離が困難です"
    if corr_max >= 0.85:
        return "🟡中", "他施策との同時実施が多く解釈注意です"
    if abs_std_coef < 0.01:
        return "🟡中", "影響度は小さめです"
    return "🟢高", "比較的評価しやすい施策です"

def confidence_score(label):
    """おすすめスコア用に、信頼度を数値化する。"""
    if str(label).startswith("🟢"):
        return 1.0
    if str(label).startswith("🟡"):
        return 0.6
    return 0.25


def star_rating(score, max_score):
    """おすすめスコアを★1〜5に変換する。"""
    if max_score <= 0 or score <= 0:
        return "☆☆☆☆☆"
    ratio = score / max_score
    if ratio >= 0.80:
        return "★★★★★"
    if ratio >= 0.60:
        return "★★★★☆"
    if ratio >= 0.40:
        return "★★★☆☆"
    if ratio >= 0.20:
        return "★★☆☆☆"
    return "★☆☆☆☆"


def make_action_comment(row):
    """営業向けの短い診断コメント。"""
    if row["推定CV差分_0to1"] <= 0:
        return "CV増加方向の効果は弱め/マイナス推定です。優先度は低めです。"
    if str(row["信頼度"]).startswith("🔴"):
        return "効果は大きく見えますが、データ条件が弱いため参考値です。追加データで再確認推奨です。"
    if str(row["信頼度"]).startswith("🟡"):
        return "プラス推定ですが、同時実施やON日数の偏りがあるため解釈注意です。"
    return "プラス推定かつ比較的評価しやすい施策です。継続/優先候補です。"


def prepare_ridge_data(df):
    """1シート分のデータから、CVと施策フラグを取り出して学習用データを作る。"""
    if df.shape[1] <= FEATURE_START_COL_IDX:
        return None, "G列以降の施策列がありません"

    # 目的変数は cv 列優先。なければ従来どおりD列。
    if TARGET_COL_NAME in df.columns:
        y_raw = df[TARGET_COL_NAME]
        target_name = TARGET_COL_NAME
    else:
        y_raw = df.iloc[:, 3]
        target_name = str(df.columns[3])

    X_raw = df.iloc[:, FEATURE_START_COL_IDX:].copy()
    X_raw.columns = [str(c) for c in X_raw.columns]

    # 数値化。'-' や空欄は NaN 扱い。
    y = pd.to_numeric(y_raw, errors="coerce")
    X = X_raw.apply(pd.to_numeric, errors="coerce")

    # 施策列は空欄を0扱いに寄せる。ただし目的変数が欠損の行は学習から除外。
    X = X.fillna(0)
    data = pd.concat([y.rename(target_name), X], axis=1).dropna(subset=[target_name])

    if len(data) < 5:
        return None, "有効な観測数が少なすぎます"

    y_clean = data[target_name]
    X_clean = data.drop(columns=[target_name])

    # 完全に変化がない列は、係数評価の対象外にする。
    variable_cols = [c for c in X_clean.columns if X_clean[c].nunique(dropna=False) > 1]
    removed_cols = [c for c in X_clean.columns if c not in variable_cols]

    if len(variable_cols) == 0:
        return None, "変化のある施策列がありません"

    X_model = X_clean[variable_cols]

    return {
        "target_name": target_name,
        "y": y_clean,
        "X": X_model,
        "X_all": X_clean,
        "removed_cols": removed_cols,
        "n_total": len(data),
    }, None


def calc_max_corr(X):
    """各施策について、他施策との最大相関を返す。"""
    if X.shape[1] <= 1:
        return {c: 0 for c in X.columns}

    corr = X.corr().abs().fillna(0)
    result = {}
    for col in corr.columns:
        others = corr.loc[col].drop(index=col, errors="ignore")
        result[col] = float(others.max()) if len(others) else 0
    return result


def calc_correlation_pairs(X, threshold=0.70):
    """
    施策同士の相関が高い組み合わせを一覧化する。
    0/1フラグ同士の相関なので、同時にON/OFFになりやすい施策の検知に使う。
    """
    if X.shape[1] <= 1:
        return pd.DataFrame(columns=[
            "施策A", "施策B", "相関", "同時ON日数", "A_ON日数", "B_ON日数",
            "A_ON時_BもON率", "B_ON時_AもON率", "注意度", "コメント"
        ])

    corr = X.corr().fillna(0)
    rows = []
    cols = list(X.columns)
    total_days = len(X)

    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            a = cols[i]
            b = cols[j]
            corr_val = float(corr.loc[a, b])
            abs_corr = abs(corr_val)

            a_on = int((X[a] != 0).sum())
            b_on = int((X[b] != 0).sum())
            both_on = int(((X[a] != 0) & (X[b] != 0)).sum())

            a_given_b = both_on / a_on if a_on else 0
            b_given_a = both_on / b_on if b_on else 0

            if abs_corr >= 0.90:
                level = "🔴高"
                comment = "ほぼセットで動いているため、単独効果の分離はかなり困難です"
            elif abs_corr >= 0.70:
                level = "🟡中"
                comment = "同時実施傾向が強く、係数解釈には注意が必要です"
            else:
                level = "🟢低"
                comment = "強い同時実施傾向は見られません"

            if abs_corr >= threshold:
                rows.append({
                    "施策A": a,
                    "施策B": b,
                    "相関": corr_val,
                    "同時ON日数": both_on,
                    "A_ON日数": a_on,
                    "B_ON日数": b_on,
                    "A_ON時_BもON率": a_given_b,
                    "B_ON時_AもON率": b_given_a,
                    "注意度": level,
                    "コメント": comment,
                })

    if not rows:
        return pd.DataFrame(columns=[
            "施策A", "施策B", "相関", "同時ON日数", "A_ON日数", "B_ON日数",
            "A_ON時_BもON率", "B_ON時_AもON率", "注意度", "コメント"
        ])

    return (
        pd.DataFrame(rows)
        .assign(相関_abs=lambda d: d["相関"].abs())
        .sort_values("相関_abs", ascending=False)
        .drop(columns=["相関_abs"])
        .reset_index(drop=True)
    )


def calc_correlation_matrix(X):
    """施策同士の相関行列。Excel出力用。"""
    if X.shape[1] <= 1:
        return pd.DataFrame()
    return X.corr().fillna(0)


def ridge_analyze_one_sheet(df):
    prepared, error = prepare_ridge_data(df)
    if error:
        return None, error

    y = prepared["y"]
    X = prepared["X"]

    model = make_pipeline(
        StandardScaler(),
        RidgeCV(alphas=RIDGE_ALPHAS)
    )
    model.fit(X, y)

    y_pred = model.predict(X)
    rmse = float(np.sqrt(mean_squared_error(y, y_pred)))
    r2 = float(r2_score(y, y_pred))

    n = X.shape[0]
    p = X.shape[1]
    if n > p + 1:
        adj_r2 = float(1 - (1 - r2) * (n - 1) / (n - p - 1))
    else:
        adj_r2 = None

    scaler = model.named_steps["standardscaler"]
    ridge = model.named_steps["ridgecv"]

    std_coef = pd.Series(ridge.coef_, index=X.columns)
    # StandardScalerを使っているので、元スケールの係数に戻す。
    # 施策フラグが0/1なら「0→1になったときの推定CV差分」に近い見方ができる。
    raw_coef = pd.Series(ridge.coef_ / scaler.scale_, index=X.columns)

    on_days = X.sum(axis=0)
    on_rate = on_days / len(X)
    max_corr = calc_max_corr(X)

    rows = []
    for col in X.columns:
        abs_std = abs(float(std_coef[col]))
        raw = float(raw_coef[col])
        corr = max_corr.get(col, 0)
        label, comment = confidence_label(
            on_days=int(on_days[col]),
            total_days=len(X),
            abs_std_coef=abs_std,
            corr_max=corr
        )

        if raw > 0:
            direction = "プラス"
        elif raw < 0:
            direction = "マイナス"
        else:
            direction = "ほぼ影響なし"

        rows.append({
            "施策": col,
            "標準化影響度": float(std_coef[col]),
            "影響度_abs": abs_std,
            "推定CV差分_0to1": raw,
            "方向": direction,
            "ON日数": int(on_days[col]),
            "ON率": float(on_rate[col]),
            "他施策との最大相関": corr,
            "信頼度": label,
            "コメント": comment,
        })

    ranking = pd.DataFrame(rows)

    # 寄与度：係数の絶対値を全施策合計で割った比較用スコア。
    # 「何%貢献した」と断定するものではなく、モデル上の相対的な重要度として使う。
    total_abs = ranking["影響度_abs"].sum() if len(ranking) else 0
    if total_abs > 0:
        ranking["寄与度_%"] = ranking["影響度_abs"] / total_abs
    else:
        ranking["寄与度_%"] = 0

    # 平均寄与CV：推定CV差分 × ON率。
    # 施策が実際にONだった頻度込みで、期間平均にどの程度乗っていそうかの目安。
    ranking["平均寄与CV"] = ranking["推定CV差分_0to1"] * ranking["ON率"]

    # おすすめスコア：プラス方向の推定効果 × 信頼度。営業向けの優先順位作成用。
    ranking["信頼係数"] = ranking["信頼度"].apply(confidence_score)
    ranking["おすすめスコア"] = ranking["推定CV差分_0to1"].clip(lower=0) * ranking["信頼係数"]
    max_reco_score = ranking["おすすめスコア"].max() if len(ranking) else 0
    ranking["おすすめ度"] = ranking["おすすめスコア"].apply(lambda x: star_rating(float(x), float(max_reco_score)))
    ranking["診断コメント"] = ranking.apply(make_action_comment, axis=1)

    ranking = ranking.sort_values("影響度_abs", ascending=False).reset_index(drop=True)
    ranking.insert(0, "順位", ranking.index + 1)

    metrics = {
        "目的変数": prepared["target_name"],
        "観測数": n,
        "施策数": p,
        "RMSE": rmse,
        "R2": r2,
        "調整済みR2": adj_r2,
        "Best Alpha": float(ridge.alpha_),
        "除外した変化なし施策数": len(prepared["removed_cols"]),
    }

    corr_pairs = calc_correlation_pairs(X, threshold=0.70)
    corr_matrix = calc_correlation_matrix(X)

    return {
        "metrics": metrics,
        "ranking": ranking,
        "removed_cols": prepared["removed_cols"],
        "correlation_pairs": corr_pairs,
        "correlation_matrix": corr_matrix,
        "y_actual": y,
        "y_pred": y_pred,
    }, None


def write_ridge_result(ws, sheet_name, result):
    metrics = result["metrics"]
    ranking = result["ranking"]

    ws["A1"] = f"施策評価サマリー：{sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "評価指標"
    ws["A3"].font = Font(bold=True)

    row = 4
    for k, v in metrics.items():
        ws.cell(row, 1).value = k
        ws.cell(row, 2).value = v
        ws.cell(row, 1).border = thin
        ws.cell(row, 2).border = thin
        row += 1

    ws["D3"] = "読み方"
    ws["D3"].font = Font(bold=True)
    notes = [
        "R2はCVの動きをどれくらい説明できているかの目安です。",
        "RMSEは予測CVと実績CVのズレの平均的な大きさです。小さいほど良いです。",
        "標準化影響度は施策間の比較用です。絶対値が大きいほどモデル上の影響が大きいです。",
        "推定CV差分_0to1は、その施策がOFFからONになったときのCV差分の目安です。",
        "信頼度が低い施策は、係数が大きくても判断注意です。",
    ]
    for i, note in enumerate(notes, 4):
        ws.cell(i, 4).value = note

    start_row = 14
    ws.cell(start_row, 1).value = "施策ランキング"
    ws.cell(start_row, 1).font = Font(bold=True)

    headers = [
        "順位", "施策", "標準化影響度", "影響度_abs", "寄与度_%", "推定CV差分_0to1",
        "平均寄与CV", "方向", "ON日数", "ON率", "他施策との最大相関",
        "信頼度", "おすすめ度", "おすすめスコア", "コメント", "診断コメント"
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(start_row + 1, col_idx)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = thin

    for r_idx, (_, r) in enumerate(ranking.iterrows(), start_row + 2):
        values = [r[h] for h in headers]
        for c_idx, v in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.value = v
            cell.border = thin

    # 表示形式
    for r in range(start_row + 2, start_row + 2 + len(ranking)):
        ws.cell(r, 3).number_format = "0.000"      # 標準化影響度
        ws.cell(r, 4).number_format = "0.000"      # 影響度_abs
        ws.cell(r, 5).number_format = "0.0%"       # 寄与度_%
        ws.cell(r, 6).number_format = "0.000"      # 推定CV差分
        ws.cell(r, 7).number_format = "0.000"      # 平均寄与CV
        ws.cell(r, 10).number_format = "0.0%"      # ON率
        ws.cell(r, 11).number_format = "0.000"     # 最大相関
        ws.cell(r, 14).number_format = "0.000"     # おすすめスコア

    # 棒グラフ：上位10件
    if len(ranking) > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "施策影響度ランキング（絶対値）"
        chart.y_axis.title = "施策"
        chart.x_axis.title = "影響度"
        max_items = min(10, len(ranking))
        data = Reference(ws, min_col=4, min_row=start_row + 1, max_row=start_row + 1 + max_items)
        cats = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 1 + max_items)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "M14")

    # 変化なしで除外した施策
    removed_cols = result.get("removed_cols", [])
    if removed_cols:
        removed_start = start_row + 4 + len(ranking)
        ws.cell(removed_start, 1).value = "変化がないため除外した施策"
        ws.cell(removed_start, 1).font = Font(bold=True)
        for i, col_name in enumerate(removed_cols, removed_start + 1):
            ws.cell(i, 1).value = col_name

    auto_fit_columns(ws)


def write_prediction_sheet(ws, sheet_name, result):
    ws["A1"] = f"実績CVと予測CV：{sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["No", "実績CV", "予測CV", "残差"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c).value = h
        ws.cell(3, c).font = Font(bold=True)
        ws.cell(3, c).border = thin

    y_actual = result["y_actual"]
    y_pred = result["y_pred"]

    for i, (actual, pred) in enumerate(zip(y_actual, y_pred), 4):
        ws.cell(i, 1).value = i - 3
        ws.cell(i, 2).value = float(actual)
        ws.cell(i, 3).value = float(pred)
        ws.cell(i, 4).value = float(actual - pred)
        for c in range(1, 5):
            ws.cell(i, c).border = thin

    auto_fit_columns(ws)


def write_correlation_sheet(ws, sheet_name, result):
    """施策同士の相関が高い組み合わせを出力する。"""
    ws["A1"] = f"施策相関チェック：{sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "読み方"
    ws["A3"].font = Font(bold=True)
    notes = [
        "相関が高い施策ペアは、同じタイミングでON/OFFになりやすい組み合わせです。",
        "相関が高い場合、CV増減がどちらの施策によるものか分離しづらくなります。",
        "🔴高はほぼセット実施、🟡中は解釈注意の目安です。",
        "このシートは『施策の優位性ランキングをどこまで信じてよいか』の補助資料です。",
    ]
    for i, note in enumerate(notes, 4):
        ws.cell(i, 1).value = note

    pairs = result.get("correlation_pairs", pd.DataFrame())

    start_row = 10
    ws.cell(start_row, 1).value = "相関が高い施策ペア（|相関| >= 0.70）"
    ws.cell(start_row, 1).font = Font(bold=True)

    headers = [
        "施策A", "施策B", "相関", "同時ON日数", "A_ON日数", "B_ON日数",
        "A_ON時_BもON率", "B_ON時_AもON率", "注意度", "コメント"
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row + 1, c)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = thin

    if pairs.empty:
        ws.cell(start_row + 2, 1).value = "強い相関のある施策ペアは見つかりませんでした"
    else:
        for r_idx, (_, r) in enumerate(pairs.iterrows(), start_row + 2):
            for c_idx, h in enumerate(headers, 1):
                cell = ws.cell(r_idx, c_idx)
                cell.value = r[h]
                cell.border = thin

        for r in range(start_row + 2, start_row + 2 + len(pairs)):
            ws.cell(r, 3).number_format = "0.000"
            ws.cell(r, 7).number_format = "0.0%"
            ws.cell(r, 8).number_format = "0.0%"

    # 相関行列も下部に出力。施策数が多すぎる場合も、Excelの横幅に注意しつつ全部出す。
    matrix = result.get("correlation_matrix", pd.DataFrame())
    matrix_start = start_row + 5 + max(len(pairs), 1)
    ws.cell(matrix_start, 1).value = "相関行列"
    ws.cell(matrix_start, 1).font = Font(bold=True)

    if matrix.empty:
        ws.cell(matrix_start + 1, 1).value = "相関行列を作成できませんでした"
    else:
        # ヘッダー
        for c_idx, col_name in enumerate(matrix.columns, 2):
            cell = ws.cell(matrix_start + 1, c_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.border = thin

        # 行名 + 値
        for r_idx, row_name in enumerate(matrix.index, matrix_start + 2):
            ws.cell(r_idx, 1).value = row_name
            ws.cell(r_idx, 1).font = Font(bold=True)
            ws.cell(r_idx, 1).border = thin
            for c_idx, col_name in enumerate(matrix.columns, 2):
                cell = ws.cell(r_idx, c_idx)
                cell.value = float(matrix.loc[row_name, col_name])
                cell.number_format = "0.000"
                cell.border = thin

    auto_fit_columns(ws, max_width=40)


def write_diagnosis_report(ws, sheet_name, result):
    """営業・マネージャー向けの診断レポートを作る。"""
    ranking = result["ranking"].copy()
    metrics = result["metrics"]
    corr_pairs = result.get("correlation_pairs", pd.DataFrame())

    ws["A1"] = f"施策診断レポート：{sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "このシートの目的"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "CV予測モデルの結果から、優先候補・注意点・解釈上の制約を営業向けに要約します。"
    ws["A5"] = "※統計的な有意差検定ではなく、モデル上の影響度・ON日数・相関を組み合わせた実務判断用です。"

    ws["A7"] = "モデル精度"
    ws["A7"].font = Font(bold=True)
    metric_rows = [
        ("R2", metrics.get("R2")),
        ("RMSE", metrics.get("RMSE")),
        ("Best Alpha", metrics.get("Best Alpha")),
        ("観測数", metrics.get("観測数")),
        ("施策数", metrics.get("施策数")),
    ]
    for i, (k, v) in enumerate(metric_rows, 8):
        ws.cell(i, 1).value = k
        ws.cell(i, 2).value = v
        ws.cell(i, 1).border = thin
        ws.cell(i, 2).border = thin

    # 推奨候補：プラス推定かつおすすめスコア順
    reco = ranking[ranking["推定CV差分_0to1"] > 0].sort_values("おすすめスコア", ascending=False).head(10)
    start = 15
    ws.cell(start, 1).value = "優先候補ランキング"
    ws.cell(start, 1).font = Font(bold=True)
    headers = ["順位", "施策", "おすすめ度", "推定CV差分_0to1", "寄与度_%", "信頼度", "ON日数", "他施策との最大相関", "診断コメント"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start + 1, c)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = thin

    if reco.empty:
        ws.cell(start + 2, 1).value = "プラス方向で推奨できる施策は見つかりませんでした。"
    else:
        for r_idx, (_, r) in enumerate(reco.iterrows(), start + 2):
            vals = [r_idx - start - 1, r["施策"], r["おすすめ度"], r["推定CV差分_0to1"], r["寄与度_%"], r["信頼度"], r["ON日数"], r["他施策との最大相関"], r["診断コメント"]]
            for c_idx, v in enumerate(vals, 1):
                cell = ws.cell(r_idx, c_idx)
                cell.value = v
                cell.border = thin
            ws.cell(r_idx, 4).number_format = "0.000"
            ws.cell(r_idx, 5).number_format = "0.0%"
            ws.cell(r_idx, 8).number_format = "0.000"

    # 注意施策：係数は大きいが信頼度が低い/中のもの
    caution = ranking[
        (ranking["影響度_abs"] > 0) &
        (~ranking["信頼度"].astype(str).str.startswith("🟢"))
    ].sort_values("影響度_abs", ascending=False).head(10)
    caution_start = start + 5 + max(len(reco), 1)
    ws.cell(caution_start, 1).value = "解釈注意ランキング"
    ws.cell(caution_start, 1).font = Font(bold=True)
    caution_headers = ["施策", "影響度_abs", "推定CV差分_0to1", "信頼度", "ON日数", "ON率", "他施策との最大相関", "コメント"]
    for c, h in enumerate(caution_headers, 1):
        cell = ws.cell(caution_start + 1, c)
        cell.value = h
        cell.font = Font(bold=True)
        cell.border = thin

    if caution.empty:
        ws.cell(caution_start + 2, 1).value = "大きな注意対象は見つかりませんでした。"
    else:
        for r_idx, (_, r) in enumerate(caution.iterrows(), caution_start + 2):
            vals = [r["施策"], r["影響度_abs"], r["推定CV差分_0to1"], r["信頼度"], r["ON日数"], r["ON率"], r["他施策との最大相関"], r["コメント"]]
            for c_idx, v in enumerate(vals, 1):
                cell = ws.cell(r_idx, c_idx)
                cell.value = v
                cell.border = thin
            ws.cell(r_idx, 2).number_format = "0.000"
            ws.cell(r_idx, 3).number_format = "0.000"
            ws.cell(r_idx, 6).number_format = "0.0%"
            ws.cell(r_idx, 7).number_format = "0.000"

    # 相関注意サマリー
    corr_start = caution_start + 5 + max(len(caution), 1)
    ws.cell(corr_start, 1).value = "同時実施の注意"
    ws.cell(corr_start, 1).font = Font(bold=True)
    if corr_pairs.empty:
        ws.cell(corr_start + 1, 1).value = "相関0.70以上の施策ペアはありませんでした。"
    else:
        high = corr_pairs[corr_pairs["相関"].abs() >= 0.90]
        mid = corr_pairs[(corr_pairs["相関"].abs() >= 0.70) & (corr_pairs["相関"].abs() < 0.90)]
        ws.cell(corr_start + 1, 1).value = f"相関0.90以上：{len(high)}ペア / 相関0.70以上0.90未満：{len(mid)}ペア"
        ws.cell(corr_start + 2, 1).value = "相関が高い施策は、ランキング上位でも単独効果としては断定しないでください。"

    # 棒グラフ：おすすめスコア上位
    if not reco.empty:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "おすすめスコア上位"
        chart.y_axis.title = "施策"
        chart.x_axis.title = "推定CV差分×信頼度"
        max_items = min(10, len(reco))
        # 診断レポート上におすすめスコア作業列を置く
        work_col = 11
        ws.cell(start + 1, work_col).value = "おすすめスコア"
        for idx, (_, r) in enumerate(reco.head(max_items).iterrows(), start + 2):
            ws.cell(idx, work_col).value = float(r["おすすめスコア"])
        data = Reference(ws, min_col=work_col, min_row=start + 1, max_row=start + 1 + max_items)
        cats = Reference(ws, min_col=2, min_row=start + 2, max_row=start + 1 + max_items)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "K15")

    auto_fit_columns(ws, max_width=60)


def run_regression_all_sheets(input_bytes):
    xls = pd.ExcelFile(io.BytesIO(input_bytes))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    created = 0
    summary_rows = []

    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet)
        result, error = ridge_analyze_one_sheet(df)

        if error:
            summary_rows.append({
                "媒体/シート": sheet,
                "ステータス": "スキップ",
                "理由": error,
                "R2": None,
                "RMSE": None,
                "Best Alpha": None,
                "影響度トップ施策": None,
                "おすすめ施策": None,
                "おすすめスコア": None,
                "相関_高ペア数": None,
                "相関_中ペア数": None,
            })
            continue

        ws = wb.create_sheet(make_unique_sheet_name(wb, sheet[:28] + "_評価"))
        write_ridge_result(ws, sheet, result)

        pred_ws = wb.create_sheet(make_unique_sheet_name(wb, sheet[:28] + "_予測"))
        write_prediction_sheet(pred_ws, sheet, result)

        corr_ws = wb.create_sheet(make_unique_sheet_name(wb, sheet[:28] + "_相関"))
        write_correlation_sheet(corr_ws, sheet, result)

        report_ws = wb.create_sheet(make_unique_sheet_name(wb, sheet[:28] + "_診断"))
        write_diagnosis_report(report_ws, sheet, result)

        top_strategy = result["ranking"].iloc[0]["施策"] if len(result["ranking"]) else None
        corr_pairs = result.get("correlation_pairs", pd.DataFrame())
        high_corr_count = int((corr_pairs["相関"].abs() >= 0.90).sum()) if not corr_pairs.empty else 0
        mid_corr_count = int(((corr_pairs["相関"].abs() >= 0.70) & (corr_pairs["相関"].abs() < 0.90)).sum()) if not corr_pairs.empty else 0

        reco_df = result["ranking"][result["ranking"]["推定CV差分_0to1"] > 0].sort_values("おすすめスコア", ascending=False)
        recommended_strategy = reco_df.iloc[0]["施策"] if len(reco_df) else None
        recommended_score = float(reco_df.iloc[0]["おすすめスコア"]) if len(reco_df) else None

        summary_rows.append({
            "媒体/シート": sheet,
            "ステータス": "完了",
            "理由": "",
            "R2": result["metrics"]["R2"],
            "RMSE": result["metrics"]["RMSE"],
            "Best Alpha": result["metrics"]["Best Alpha"],
            "影響度トップ施策": top_strategy,
            "おすすめ施策": recommended_strategy,
            "おすすめスコア": recommended_score,
            "相関_高ペア数": high_corr_count,
            "相関_中ペア数": mid_corr_count,
        })
        created += 1

    # 全体サマリー
    summary_ws = wb.create_sheet("全体サマリー", 0)
    summary_ws["A1"] = "リッジ回帰 施策評価サマリー"
    summary_ws["A1"].font = Font(bold=True, size=14)

    if summary_rows:
        headers = list(summary_rows[0].keys())
        for c, h in enumerate(headers, 1):
            summary_ws.cell(3, c).value = h
            summary_ws.cell(3, c).font = Font(bold=True)
            summary_ws.cell(3, c).border = thin
        for r_idx, row_data in enumerate(summary_rows, 4):
            for c_idx, h in enumerate(headers, 1):
                summary_ws.cell(r_idx, c_idx).value = row_data[h]
                summary_ws.cell(r_idx, c_idx).border = thin
        auto_fit_columns(summary_ws)

    if created == 0:
        ws = wb.create_sheet("NoData")
        ws["A1"] = "有効データなし"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()

# ==============================
# UI
# ==============================
tab1, tab2 = st.tabs(["転記用生成", "回帰分析"])


with tab1:
    st.title("転記用エクセル生成ツール")

    input_file = st.file_uploader("施策一覧をアップロード", type=["xlsx"])

    if input_file:
        wb = openpyxl.load_workbook(input_file, read_only=True)
        sheet_names = wb.sheetnames

        selected_sheets = st.multiselect(
            "対象シートを選択（複数選択可）",
            sheet_names,
            default=sheet_names[:1]
        )

        st.caption("選択した元データシートをまとめて読み込み、出力ファイルは従来どおり媒体ごとにシート分けします。")

        if st.button("実行"):
            if not selected_sheets:
                st.warning("対象シートを1つ以上選択してください。")
            else:
                with open(FMT_PATH, "rb") as f:
                    fmt_bytes = f.read()

                output, sheet_count, total_days = build_workbook(
                    input_file.getvalue(),
                    fmt_bytes,
                    selected_sheets
                )

                today_str = datetime.now().strftime("%Y%m%d")

                st.success(f"✅ 完成！ 選択シート:{len(selected_sheets)} 媒体:{sheet_count} 日数:{total_days}")
                st.balloons()

                st.download_button(
                    "ダウンロード",
                    data=output,
                    file_name=f"転記用_{today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


with tab2:
    st.title("回帰分析ツール")

    reg_file = st.file_uploader("回帰分析用ファイルをアップロード", type=["xlsx"], key="reg")

    if reg_file:
        if st.button("回帰分析実行"):

            result = run_regression_all_sheets(reg_file.getvalue())
            today_str = datetime.now().strftime("%Y%m%d")

            st.success("✅ 回帰分析完了！")
            st.balloons()

            st.download_button(
                "ダウンロード",
                data=result,
                file_name=f"回帰分析結果_{today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
